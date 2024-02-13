from rest_framework.exceptions import AuthenticationFailed
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.conf import settings

import pyexasol

from sqlalchemy import create_engine, MetaData, text, Table, inspect, insert
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.exc import IntegrityError
from sqlalchemy.engine.reflection import Inspector

from django.middleware.csrf import get_token
from django.http import JsonResponse
import pandas as pd
from rest_framework import status
from datetime import datetime
import time
import openpyxl


# EXASOL_CONNECTION_STRING = "exa+pyodbc://EXADB"
EXASOL_CONNECTION_STRING = settings.EXASOL_CONNECTION_STRING


@api_view(['POST'])
@permission_classes([])
def excel_uploader_v2(request):
    start_time = time.time()
    print('Hello, Im in excel_uploader')
    new_data = request.data.get('data', [])
    file_name = request.data.get('file')
    print(file_name)
    schema_name = request.data.get('schema')
    table_name = request.data.get('table')
    sheet_name = request.data.get('sheet')
    username = request.data.get('username')

    print(schema_name, table_name, sheet_name)

    default_values = {
        "SOURCE_FILE_NAME": str(file_name),
        "SHEET_NAME": str(sheet_name),
        "USER_NAME": username,
        "LOAD_TIMESTAMP": datetime.now()
    }

    try:
        excel_columns = get_excel_columns(request.FILES['file'], sheet_name)
        print('#' * 10)
        print('excel columns')
        print(excel_columns)
        exasol_columns = get_exasol_columns(schema_name, table_name)
        print('*' * 10)
        print('EXASOL columns')
        print(exasol_columns)

        missing_elements = [item for item in excel_columns if item not in exasol_columns]


        if not missing_elements:
            print('All elements are okay')
        else:
            print('These elements are:', missing_elements)

        excel_set = set(excel_columns)
        exasol_set = set(exasol_columns)

        if excel_set.issubset(exasol_set):
            print('Data is successfully checked')
            start_row = 2
            chunk_size = 10000
            while True:
                data_to_send = read_data_from_excel(file_name, sheet_name, start_row, excel_columns, chunk_size)
                if not data_to_send:
                    break
                create_data_in_exasol(data_to_send, table_name, schema_name, exasol_columns, default_values, file_name)
                start_row += chunk_size
            end_time = time.time()
            execution_time_overall = end_time - start_time
            response = {'message': f'Данные успешно выгружены. Затраченное время: {execution_time_overall}'}
            print('Printing response on create:', response)
            print(f'Execution time: {execution_time_overall}')
        else:
            response = {
                'error': 'Название полей Excel и Базы данных не соответствуют друг другу.'}
            print('Название полей Excel и Базы данных не соответствуют друг другу.')

    except Exception as e:
        response = {'error': str(e)}
        print('Error in create operation:', response)
    

    return Response(response)

def get_exasol_columns(schema_name, table_name):
    print('')
    print('Im in get_exasol_columns')
    start = time.time()
    print(schema_name)
    print(table_name)
    engine = create_engine(EXASOL_CONNECTION_STRING)
    # connection = engine.connect()
    metadata = MetaData(bind=engine)

    table = Table(f"{table_name}", metadata, autoload=True, schema=schema_name)
 
    columns = table.columns.keys()

    columns = table.columns.keys()
    columns = [x.upper() for x in columns]

    end = time.time()

    overall_time = end - start
    print(f'get_exasol_columns time execution: {overall_time}')

    # sql_statement = text(
    #     """
    #     SELECT COLUMN_NAME
    #     FROM EXA_ALL_COLUMNS
    #     WHERE COLUMN_SCHEMA = :schema_name
    #     AND COLUMN_TABLE = :table_name
    #     """
    # )

    # result = connection.execute(
    #     sql_statement, schema_name=schema_name, table_name=table_name).fetchall()

    # columns = [row[0] for row in result]

    # connection.close()

    return columns


def get_excel_columns(file_path, sheet_name):
    print('')
    print('get_excel_columns started')
    start = time.time()

    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    
    columns = [cell.value for cell in sheet[1]]
    
    end = time.time()
    overall_time = end - start
    print(f'get_excel_columns time execution: {overall_time}')

    return columns


def read_data_from_excel(file_path, sheet_name, start_row, excel_columns, chunk_size):
    print('')
    print('read_data_from_excel started')
    start = time.time()

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[sheet_name]
    data_rows = list(sheet.iter_rows(min_row=start_row, max_row=start_row + chunk_size - 1, values_only=True))
    data = []

    for row in data_rows:
        data.append({key.lower(): value for key, value in zip(excel_columns, row)})
    
    
    end = time.time()
    overall_time = end - start
    print(f'read_data_from_excel time execution: {overall_time}')


    return data


def create_data_in_exasol(data, table_name, schema_name, exasol_columns, default_values, file_name):
    print("")
    print("create_data_in_exasol started")
    start_data_send = time.time()

    engine = create_engine(EXASOL_CONNECTION_STRING)

    Session = sessionmaker(bind=engine)
    
    # columns = exasol_columns
    with Session() as session:
        metadata = MetaData(bind=engine)
        # column_names = ", ".join(f'"{col.strip()}"' for col in columns)
        # value_placeholders = ", ".join(f':{col.strip().replace(" ", "_")}' for col in columns)
        column_names = [col.strip() for col in exasol_columns]
        # table = Table(table_name, metadata, autoload=True, schema=schema_name)

        with engine.connect() as connection:
            trans = connection.begin()

            try:
                rows_to_insert = []
                dv_ins_start = time.time()
                for row in data:
                    for key, value in default_values.items():
                        row[key.lower()] = str(value)
                    # row["source_file_name"] = "test_source_name"
                    # row["sheet_name"] = "test_sheet_name"
                    # row["user_name"] = "test_user_name"
                    # row["load_timestamp"] = datetime.now
                    rows_to_insert.append(row)
                wb = openpyxl.Worbook()
                ws = wb.activate

                columns = exasol_columns
                for col_index, col_name in enumerate(columns, start=1):
                    ws.cell(row=1, column=cold_index, value=col_name)

                for row_index, row_data in enumerate(rows_to_insert, start=2):
                    for col_index, col_name in eumerate(columns, start=1):
                        ws.cell(row=row_index, column=col_index, value=row_data.get(col_name, ""))
                


                dv_ins_end = time.time()
                dv_ins_result = dv_ins_end - dv_ins_start
                print("")
                print(f"inserting def values took: {dv_ins_result}")
                print(file_name)
                print("")
                
                # if len(rows_to_insert) > 0:
                #     stmt = insert(table)
                #     connection.execute(stmt, rows_to_insert)

                sql_query = f"""
                    IMPORT INTO "{schema_name}."{table_name}"
                    ({column_names})
                    FROM CSV FILE "{file_name}"
                """

                connection.execute(sql_query)
                trans.commit()

            except Exception as e:
                trans.rollback()
                print(f"Error: {e}")
                raise

    engine.dispose()

    end_data_send = time.time()
    overall_data_send = end_data_send - start_data_send 
    print(f'overall_data_send: {overall_data_send}')



@api_view(['POST'])
@permission_classes([])
def retrieve_sheet_names(request):
    print('Im in retreive_sheet_names func')
    try:
        uploaded_file = request.FILES.get('excelFile')
        if not uploaded_file:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        excel_data = pd.ExcelFile(uploaded_file)
        sheet_names = excel_data.sheet_names

        return Response({'sheetNames': sheet_names}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# -- DISTINCT GL_ACCOUNT_DT AS ACC_NUMBER WITH INVALID FOREIGN KEY (ACC_NUMBER) FROM AMS_GL.XLSX_DGL_BUFFER REFERENCING AMS_GL.GL_ACCOUNT;
# --SELECT DISTINCT "Линия Бизнеса ID" WITH INVALID FOREIGN KEY ("Линия Бизнеса ID") FROM T2_MOTORSALES_NEW."Полис" REFERENCING IDM_MOTORSALES_NEW."Линия Бизнеса";
# INSERT INTO AMS_GL.XLSX_DGL_BUFFER_ERRORS 
# (
# 	SOURCE_FILE_NAME, 
# 	SHEET_NAME, 
# 	USER_NAME, 
# 	LINE_ID, 
# 	FIELD_NAME, 
# 	ERROR_VALUE, 
# 	ERROR_TYPE
# )
# SELECT DISTINCT SOURCE_FILE_NAME, SHEET_NAME, USER_NAME, LINE_ID, 'GL_ACCOUNT_DT' AS FIELD_NAME,GL_ACCOUNT_DT AS ERROR_VALUE, 'Значение поля не найдено в справочнике' AS ERROR_TYPE   WITH INVALID FOREIGN KEY (GL_ACCOUNT_DT) FROM AMS_GL.XLSX_DGL_BUFFER REFERENCING AMS_GL.GL_ACCOUNT (ACC_NUMBER) WHERE GL_ACCOUNT_DT IS NOT NULL;